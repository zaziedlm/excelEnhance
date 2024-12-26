import os
import glob
import sys
from openpyxl import load_workbook
# dotenvの読み込み
from dotenv import load_dotenv

# .envファイルの読み込み
load_dotenv()

# Function to extract non-empty cells from a sheet
def extract_non_empty_cells(sheet):
    non_empty_cells = {}
    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.value:  # Only include non-empty cells
                non_empty_cells[cell.coordinate] = cell.value
    return non_empty_cells

def annotate_sheet_structure(workbook):
    """
    ワークブック内のシート構成をまとめるサマリ情報を生成する。
    """
    summary = "---WorkbookSheetStructure---\n"
    for idx, sheet_name in enumerate(workbook.sheetnames, start=1):
        summary += f"SheetIndex={idx};SheetName={sheet_name}\n"
    return summary

# Function to generate Markdown content for an Excel file
def generate_markdown_from_excel(file_path):
    workbook = load_workbook(file_path)
    file_name = file_path.split('/')[-1]
    markdown_output = f"# ファイル名: {file_name}\n\n"
    
    # シート構成サマリを追加
    markdown_output += "## シート構成サマリ\n"
    markdown_output += "```markdown\n"
    markdown_output += annotate_sheet_structure(workbook)
    markdown_output += "```\n\n"
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        non_empty_cells = extract_non_empty_cells(sheet)
        
        markdown_output += f"## {sheet_name}\n"
        markdown_output += "```markdown\n"
        
        for cell, value in non_empty_cells.items():
            markdown_output += f"### {cell}: {value}\n"
        
        markdown_output += "```\n\n"
    
    return markdown_output


# 環境変数から入出力ディレクトリを取得（なければデフォルト ./input, ./output を使用）
input_dir = os.getenv("FILES_IN_DIR", "./data-input")
output_dir = os.getenv("FILES_OUTMD_DIR", "./data-outmd")

# 出力先ディレクトリが存在しない場合は作成しておく
os.makedirs(output_dir, exist_ok=True)

# input_dir 内の .xlsx ファイルをすべて取得
xlsx_files = glob.glob(os.path.join(input_dir, "*.xlsx"))

if not xlsx_files:
    print(f"[Warning] No .xlsx files found in {input_dir}")
    sys.exit(0)

# ファイルごとに処理
for file_path in xlsx_files:
    # Usage example
    markdown_result = generate_markdown_from_excel(file_path)

    # Save the result to a Markdown file (optional)
    base_name = os.path.basename(file_path)  # 例: "sample.xlsx"
    md_name = os.path.splitext(base_name)[0] + ".md"  # 拡張子を .md に変更
    output_file = os.path.join(output_dir, md_name)  # 同じファイル名で出力
    print(f"Processing: {file_path} -> {output_file}")
    
    with open(output_file, "w", encoding="utf-8") as f:
        f.write(markdown_result)

    # Print the Markdown result to the console
    print(markdown_result)
